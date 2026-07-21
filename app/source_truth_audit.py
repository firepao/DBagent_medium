"""Read-only consistency audit for source workbooks, SQLite, DDL, and table cards."""

from __future__ import annotations

import hashlib
import json
import sqlite3
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


@dataclass(frozen=True)
class AuditIssue:
    code: str
    severity: str
    message: str

    def as_dict(self) -> dict[str, str]:
        return {"code": self.code, "severity": self.severity, "message": self.message}


class SourceTruthAuditor:
    """Compare published query metadata with the actual imported SQLite database.

    The auditor never opens a writable SQLite connection and never changes source
    workbooks. It treats SQLite as the runtime schema fact and reports all drift.
    """

    def __init__(
        self,
        *,
        source_data_dir: str | Path,
        processed_data_dir: str | Path,
        sqlite_db_path: str | Path,
        ddl_directory: str | Path,
        catalog_path: str | Path,
        table_cards_path: str | Path,
        ddl_registry_path: str | Path,
    ) -> None:
        self.source_data_dir = Path(source_data_dir)
        self.processed_data_dir = Path(processed_data_dir)
        self.sqlite_db_path = Path(sqlite_db_path)
        self.ddl_directory = Path(ddl_directory)
        self.catalog = self._load_json(catalog_path)
        self.table_cards = {
            card["table"]: card
            for card in self._load_json(table_cards_path).get("table_cards", [])
            if card.get("table")
        }
        self.ddl_registry = self._load_json(ddl_registry_path).get("tables", {})
        self.datasets = {
            dataset["table"]: dataset
            for dataset in self.catalog.get("datasets", [])
            if dataset.get("table")
        }

    @staticmethod
    def _load_json(path: str | Path) -> dict[str, Any]:
        return json.loads(Path(path).read_text(encoding="utf-8"))

    def run(self) -> dict[str, Any]:
        if not self.sqlite_db_path.is_file():
            raise FileNotFoundError(f"SQLite 数据库不存在：{self.sqlite_db_path}")

        tables = [self._audit_table(table) for table in sorted(self.datasets)]
        failed = sum(item["status"] == "fail" for item in tables)
        warned = sum(item["status"] == "warning" for item in tables)
        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_data_dir": str(self.source_data_dir),
            "processed_data_dir": str(self.processed_data_dir),
            "sqlite_db_path": str(self.sqlite_db_path),
            "summary": {
                "table_count": len(tables),
                "passed_tables": len(tables) - failed - warned,
                "warning_tables": warned,
                "failed_tables": failed,
            },
            "tables": tables,
        }

    def _audit_table(self, table: str) -> dict[str, Any]:
        issues: list[AuditIssue] = []
        dataset = self.datasets[table]
        table_exists = self._table_exists(table)
        sqlite_columns = self._table_columns(table) if table_exists else []
        row_count = self._row_count(table) if table_exists else None

        if not table_exists:
            issues.append(AuditIssue("SQLITE_TABLE_MISSING", "error", f"SQLite 中不存在表 {table}"))

        source_file, source_sheet = self._source_trace(table)
        raw_sheet_name = self._raw_sheet_mapping(table, source_sheet)
        raw_path = self._find_source_file(self.source_data_dir, source_file)
        processed_path = self._find_source_file(self.processed_data_dir, source_file)
        raw_exists = raw_path is not None
        processed_exists = processed_path is not None
        raw_sheet_exists, raw_sheet_summary, raw_sheet_error = self._read_sheet_summary(
            raw_path, raw_sheet_name
        )
        processed_sheet_exists, processed_sheet_summary, processed_sheet_error = self._read_sheet_summary(
            processed_path, source_sheet
        )
        raw_sheet_status = self._sheet_status(raw_exists, raw_sheet_name, raw_sheet_exists, raw_sheet_error)
        if source_file:
            if not raw_exists:
                issues.append(AuditIssue("RAW_SOURCE_FILE_MISSING", "error", f"原始数据目录未找到 {source_file}"))
            if not processed_exists:
                issues.append(AuditIssue("PROCESSED_SOURCE_FILE_MISSING", "error", f"处理后数据目录未找到 {source_file}"))
        else:
            issues.append(AuditIssue("SOURCE_TRACEABILITY_MISSING", "warning", "未在导入元数据中找到来源文件"))

        self._append_sheet_issue(issues, "RAW", raw_exists, raw_sheet_exists, raw_sheet_error)
        self._append_sheet_issue(
            issues, "PROCESSED", processed_exists, processed_sheet_exists, processed_sheet_error
        )

        public_columns = set(sqlite_columns) - set(dataset.get("excluded_columns", []))
        card = self.table_cards.get(table)
        if card is None:
            issues.append(AuditIssue("TABLE_CARD_MISSING", "error", "缺少 TableCard 配置"))
        else:
            for term, field in card.get("aliases", {}).items():
                if field not in public_columns:
                    issues.append(
                        AuditIssue(
                            "TABLE_CARD_ALIAS_FIELD_MISSING",
                            "error",
                            f"TableCard aliases 中“{term}”引用的字段 {field} 不在 SQLite 已发布字段中",
                        )
                    )
            for field in card.get("important_fields", []):
                if field not in public_columns:
                    issues.append(
                        AuditIssue(
                            "TABLE_CARD_IMPORTANT_FIELD_MISSING",
                            "error",
                            f"TableCard important_fields 引用的字段 {field} 不在 SQLite 已发布字段中",
                        )
                    )

        for field in dataset.get("aliases", {}):
            if field not in public_columns:
                issues.append(
                    AuditIssue(
                        "CATALOG_ALIAS_FIELD_MISSING",
                        "error",
                        f"catalog aliases 引用的字段 {field} 不在 SQLite 已发布字段中",
                    )
                )

        ddl_info, ddl_issues = self._audit_ddl(table, sqlite_columns)
        issues.extend(ddl_issues)
        vanna_table, vanna_fields, import_audit, metadata_issues = self._audit_import_metadata(
            table, row_count
        )
        issues.extend(metadata_issues)

        status = "fail" if any(issue.severity == "error" for issue in issues) else (
            "warning" if issues else "pass"
        )
        return {
            "table": table,
            "status": status,
            "source": {
                "source_file": source_file,
                "source_sheet": source_sheet,
                "raw_exists": raw_exists,
                "processed_exists": processed_exists,
                "raw_sheet_mapping": raw_sheet_name,
                "raw_sheet_exists": raw_sheet_exists,
                "raw_sheet_status": raw_sheet_status,
                "processed_sheet_exists": processed_sheet_exists,
                "raw_sheet_summary": raw_sheet_summary,
                "processed_sheet_summary": processed_sheet_summary,
            },
            "sqlite": {
                "exists": table_exists,
                "columns": sqlite_columns,
                "published_columns": sorted(public_columns),
                "row_count": row_count,
                "vanna_table": vanna_table,
                "vanna_fields": vanna_fields,
                "import_audit": import_audit,
            },
            "ddl": ddl_info,
            "table_card": {
                "exists": card is not None,
                "aliases": (card or {}).get("aliases", {}),
                "important_fields": (card or {}).get("important_fields", []),
            },
            "issues": [issue.as_dict() for issue in issues],
        }

    def _connect(self) -> sqlite3.Connection:
        uri = f"file:{self.sqlite_db_path.resolve().as_posix()}?mode=ro"
        return sqlite3.connect(uri, uri=True)

    @staticmethod
    def _ident(value: str) -> str:
        return '"' + value.replace('"', '""') + '"'

    def _table_exists(self, table: str) -> bool:
        with self._connect() as connection:
            return connection.execute(
                "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?", (table,)
            ).fetchone() is not None

    def _table_columns(self, table: str) -> list[str]:
        with self._connect() as connection:
            return [row[1] for row in connection.execute(f"PRAGMA table_info({self._ident(table)})")]

    def _row_count(self, table: str) -> int:
        with self._connect() as connection:
            return int(connection.execute(f"SELECT COUNT(*) FROM {self._ident(table)}").fetchone()[0])

    def _has_table(self, connection: sqlite3.Connection, table: str) -> bool:
        return connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?", (table,)
        ).fetchone() is not None

    def _source_trace(self, table: str) -> tuple[str | None, str | None]:
        with self._connect() as connection:
            if self._has_table(connection, "vanna_tables"):
                row = connection.execute(
                    "SELECT source_file, source_sheet FROM vanna_tables WHERE table_name = ?", (table,)
                ).fetchone()
                if row and row[0]:
                    return str(row[0]), str(row[1] or "")
            if self._has_table(connection, "import_audit"):
                row = connection.execute(
                    "SELECT source_file, source_sheet FROM import_audit WHERE table_name = ? ORDER BY rowid DESC LIMIT 1",
                    (table,),
                ).fetchone()
                if row and row[0]:
                    return str(row[0]), str(row[1] or "")
            columns = self._table_columns(table)
            if "source_file" in columns:
                row = connection.execute(
                    f"SELECT source_file, source_sheet FROM {self._ident(table)} WHERE source_file IS NOT NULL LIMIT 1"
                ).fetchone()
                if row and row[0]:
                    return str(row[0]), str(row[1] or "")
        return None, None

    def _raw_sheet_mapping(self, table: str, processed_sheet: str | None) -> str | None:
        dataset = self.datasets[table]
        direct = str(dataset.get("raw_source_sheet") or "").strip()
        if direct:
            return direct
        mapping = dataset.get("raw_sheet_mapping")
        if isinstance(mapping, dict) and processed_sheet:
            value = str(mapping.get(processed_sheet) or "").strip()
            return value or None
        return None

    @staticmethod
    def _sheet_status(
        workbook_exists: bool,
        mapped_sheet: str | None,
        sheet_exists: bool | None,
        error: str | None,
    ) -> str:
        if not workbook_exists:
            return "file_missing"
        if not mapped_sheet:
            return "not_mapped"
        if error:
            return "unreadable"
        return "verified" if sheet_exists else "missing"

    @staticmethod
    def _find_source_file(directory: Path, source_file: str | None) -> Path | None:
        if not source_file:
            return None
        direct = directory / source_file
        if direct.is_file():
            return direct
        return next(
            (candidate for candidate in directory.rglob(Path(source_file).name) if candidate.is_file()),
            None,
        )

    @staticmethod
    def _append_sheet_issue(
        issues: list[AuditIssue],
        prefix: str,
        workbook_exists: bool,
        sheet_exists: bool | None,
        error: str | None,
    ) -> None:
        if not workbook_exists:
            return
        if error:
            issues.append(AuditIssue(f"{prefix}_SOURCE_WORKBOOK_UNREADABLE", "error", error))
        elif sheet_exists is False:
            issues.append(
                AuditIssue(
                    f"{prefix}_SOURCE_SHEET_MISSING",
                    "error",
                    "Registered source sheet does not exist in the workbook.",
                )
            )

    @classmethod
    def _read_sheet_summary(
        cls, workbook_path: Path | None, sheet_name: str | None
    ) -> tuple[bool | None, dict[str, Any] | None, str | None]:
        """Read workbook metadata without persisting source cell values."""
        if workbook_path is None or not sheet_name:
            return None, None, None
        if workbook_path.suffix.lower() != ".xlsx":
            return None, None, "Only .xlsx source workbooks are supported."

        try:
            try:
                return cls._read_sheet_summary_openpyxl(workbook_path, sheet_name)
            except ModuleNotFoundError:
                return cls._read_sheet_summary_xlsx_xml(workbook_path, sheet_name)
        except (OSError, ValueError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
            return None, None, f"Unable to read source workbook: {exc}"

    @staticmethod
    def _summarize_rows(rows: list[list[Any]]) -> dict[str, Any] | None:
        non_empty_rows = [
            [str(value).strip() for value in row if value is not None and str(value).strip()]
            for row in rows
        ]
        non_empty_rows = [row for row in non_empty_rows if row]
        if not non_empty_rows:
            return None
        header = non_empty_rows[0]
        sample = next((row for row in non_empty_rows[1:] if row), [])
        return {
            "headers": header,
            "sample_row_count": 1 if sample else 0,
            "sample_non_empty_cell_count": len(sample),
        }

    @classmethod
    def _read_sheet_summary_openpyxl(
        cls, workbook_path: Path, sheet_name: str
    ) -> tuple[bool, dict[str, Any] | None, None]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                return False, None, None
            worksheet = workbook[sheet_name]
            rows = [list(row) for row in worksheet.iter_rows(max_row=6, values_only=True)]
            return True, cls._summarize_rows(rows), None
        finally:
            workbook.close()

    @classmethod
    def _read_sheet_summary_xlsx_xml(
        cls, workbook_path: Path, sheet_name: str
    ) -> tuple[bool, dict[str, Any] | None, None]:
        with zipfile.ZipFile(workbook_path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            targets = {
                relation.attrib["Id"]: relation.attrib["Target"]
                for relation in relationships
                if relation.tag.endswith("Relationship")
            }
            selected = next(
                (
                    sheet
                    for sheet in workbook.iter()
                    if sheet.tag.endswith("sheet") and sheet.attrib.get("name") == sheet_name
                ),
                None,
            )
            if selected is None:
                return False, None, None
            relation_id = next(
                (value for key, value in selected.attrib.items() if key.endswith("}id")), None
            )
            target = targets.get(relation_id or "")
            if not target:
                raise ValueError("Worksheet relationship is missing.")
            sheet_path = target.lstrip("/")
            if not sheet_path.startswith("xl/"):
                sheet_path = f"xl/{sheet_path}"
            sheet = ElementTree.fromstring(archive.read(sheet_path))
            rows = [
                cls._read_xml_row(row, cls._read_shared_strings(archive))
                for row in sheet.iter()
                if row.tag.endswith("row")
            ][:6]
            return True, cls._summarize_rows(rows), None

    @staticmethod
    def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
        try:
            root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
        except KeyError:
            return []
        return [
            "".join(node.text or "" for node in item.iter() if node.tag.endswith("t"))
            for item in root
            if item.tag.endswith("si")
        ]

    @staticmethod
    def _read_xml_row(row: ElementTree.Element, shared_strings: list[str]) -> list[str]:
        values: list[str] = []
        for cell in row:
            if not cell.tag.endswith("c"):
                continue
            cell_type = cell.attrib.get("t")
            text_nodes = [node.text or "" for node in cell.iter() if node.tag.endswith("t")]
            value_node = next((node.text for node in cell if node.tag.endswith("v")), None)
            if cell_type == "s" and value_node is not None:
                index = int(value_node)
                value = shared_strings[index] if index < len(shared_strings) else ""
            elif text_nodes:
                value = "".join(text_nodes)
            else:
                value = value_node or ""
            values.append(value)
        return values

    def _audit_ddl(self, table: str, sqlite_columns: list[str]) -> tuple[dict[str, Any], list[AuditIssue]]:
        issues: list[AuditIssue] = []
        entry = self.ddl_registry.get(table)
        if not entry:
            return {"registered": False}, [AuditIssue("DDL_REGISTRY_MISSING", "error", "DDL 注册缺失")]
        path = self.ddl_directory / str(entry.get("file", ""))
        info: dict[str, Any] = {"registered": True, "file": str(path), "exists": path.is_file()}
        if not path.is_file():
            return info, [AuditIssue("DDL_FILE_MISSING", "error", f"DDL 文件不存在：{path.name}")]

        ddl = path.read_text(encoding="utf-8")
        digest = hashlib.sha256(ddl.encode("utf-8")).hexdigest()
        info["sha256"] = digest
        if entry.get("sha256") and entry["sha256"] != digest:
            issues.append(AuditIssue("DDL_SHA256_MISMATCH", "error", "DDL 文件哈希与注册表不一致"))
        try:
            with sqlite3.connect(":memory:") as connection:
                connection.executescript(ddl)
                ddl_signature = self._schema_signature(connection, table)
        except sqlite3.DatabaseError as exc:
            issues.append(AuditIssue("DDL_PARSE_FAILED", "error", f"DDL 无法解析：{exc}"))
            info["columns"] = []
            return info, issues
        info["columns"] = [column[0] for column in ddl_signature]
        with self._connect() as connection:
            sqlite_signature = self._schema_signature(connection, table)
        if ddl_signature != sqlite_signature:
            issues.append(
                AuditIssue(
                    "DDL_SQLITE_STRUCTURE_MISMATCH",
                    "error",
                    "DDL 列顺序、字段类型、非空或主键属性与 SQLite 不一致",
                )
            )
        return info, issues

    def _schema_signature(
        self, connection: sqlite3.Connection, table: str
    ) -> list[tuple[str, str, int, int]]:
        return [
            (row[1], (row[2] or "").upper(), int(row[3]), int(row[5]))
            for row in connection.execute(f"PRAGMA table_info({self._ident(table)})")
        ]

    def _audit_import_metadata(
        self, table: str, row_count: int | None
    ) -> tuple[dict[str, Any] | None, list[str], list[dict[str, Any]], list[AuditIssue]]:
        issues: list[AuditIssue] = []
        vanna_table: dict[str, Any] | None = None
        vanna_fields: list[str] = []
        import_audit: list[dict[str, Any]] = []
        with self._connect() as connection:
            if self._has_table(connection, "vanna_tables"):
                row = connection.execute(
                    "SELECT source_file, source_sheet, row_count FROM vanna_tables WHERE table_name = ?", (table,)
                ).fetchone()
                if row:
                    vanna_table = {"source_file": row[0], "source_sheet": row[1], "row_count": row[2]}
                    if row_count is not None and row[2] is not None and int(row[2]) != row_count:
                        issues.append(AuditIssue("VANNA_TABLE_ROW_COUNT_MISMATCH", "warning", "vanna_tables 行数与 SQLite 不一致"))
                else:
                    issues.append(AuditIssue("VANNA_TABLE_METADATA_MISSING", "warning", "vanna_tables 缺少该表元数据"))
            if self._has_table(connection, "vanna_fields"):
                vanna_field_columns = {
                    row[1] for row in connection.execute("PRAGMA table_info(vanna_fields)")
                }
                order_by = "ordinal, column_name" if "ordinal" in vanna_field_columns else "column_name"
                vanna_fields = [
                    str(row[0])
                    for row in connection.execute(
                        f"SELECT column_name FROM vanna_fields WHERE table_name = ? ORDER BY {order_by}", (table,)
                    )
                ]
                excluded_fields = set(self.datasets[table].get("excluded_columns", []))
                sqlite_business_fields = set(self._table_columns(table)) - excluded_fields
                vanna_business_fields = set(vanna_fields) - excluded_fields
                if vanna_business_fields != sqlite_business_fields:
                    issues.append(
                        AuditIssue(
                            "VANNA_FIELDS_SCHEMA_MISMATCH",
                            "warning",
                            "vanna_fields 字段集合与 SQLite 实际字段集合不一致",
                        )
                    )
            if self._has_table(connection, "import_audit"):
                columns = [row[1] for row in connection.execute("PRAGMA table_info(import_audit)")]
                if {"table_name", "source_file", "source_sheet"}.issubset(columns):
                    rows = connection.execute(
                        "SELECT source_file, source_sheet, row_count FROM import_audit WHERE table_name = ? ORDER BY rowid", (table,)
                    ).fetchall()
                    import_audit = [
                        {"source_file": row[0], "source_sheet": row[1], "row_count": row[2]} for row in rows
                    ]
        return vanna_table, vanna_fields, import_audit, issues


def render_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        "# 源数据事实核查报告",
        "",
        f"- 核查表数：{summary['table_count']}",
        f"- 通过：{summary['passed_tables']}；警告：{summary['warning_tables']}；失败：{summary['failed_tables']}",
        "",
        "| 表 | 状态 | SQLite 行数 | 原始数据 | 处理后数据 | 问题 |",
        "| --- | --- | ---: | --- | --- | --- |",
    ]
    for item in report["tables"]:
        issues = "<br>".join(issue["code"] for issue in item["issues"]) or "无"
        source = item["source"]
        lines.append(
            "| {table} | {status} | {rows} | {raw} | {processed} | {issues} |".format(
                table=item["table"],
                status=item["status"],
                rows=item["sqlite"]["row_count"] if item["sqlite"]["row_count"] is not None else "-",
                raw="存在" if source["raw_exists"] else "缺失/未知",
                processed="存在" if source["processed_exists"] else "缺失/未知",
                issues=issues,
            )
        )
    lines.extend(["", "## 逐表问题", ""])
    for item in report["tables"]:
        if not item["issues"]:
            continue
        lines.extend([f"### {item['table']}", ""])
        lines.extend(f"- `{issue['code']}` ({issue['severity']})：{issue['message']}" for issue in item["issues"])
        lines.append("")
    return "\n".join(lines)
